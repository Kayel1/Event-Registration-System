from flask import Flask, render_template, request, redirect, url_for, flash, session
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os
from werkzeug.security import check_password_hash, generate_password_hash
from config import config, ADMIN_USERNAME, ADMIN_PASSWORD, REGISTRATIONS_FILE, EVENTS_FILE, SECRET_KEY
from functools import wraps
from time import time
import re
import threading
import logging
from logging.handlers import RotatingFileHandler
from blueprints.admin import admin
from forms import RegistrationForm

# Constants
MAX_ATTEMPTS = 3
LOCKOUT_TIME = 300  # 5 minutes in seconds

# Create Flask app
app = Flask(__name__)
env = os.environ.get('FLASK_ENV', 'default')
app.config.from_object(config[env])

# Ensure secret key is set
if not app.secret_key:
    app.secret_key = SECRET_KEY

# Register blueprints
app.register_blueprint(admin)

# Initialize CSRF protection
csrf = CSRFProtect(app)

# Initialize rate limiter
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["100 per hour"],
    storage_uri="memory://"
)

# Configure logging
if not os.path.exists(app.config['LOG_FOLDER']):
    os.makedirs(app.config['LOG_FOLDER'])

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('event_registration')
logger.setLevel(logging.INFO)

# File handler with improved configuration
file_handler = RotatingFileHandler(
    os.path.join(app.config['LOG_FOLDER'], 'app.log'),
    maxBytes=app.config['LOG_MAX_BYTES'],
    backupCount=app.config['LOG_BACKUP_COUNT']
)
file_handler.setFormatter(logging.Formatter(app.config['LOG_FORMAT']))
logger.addHandler(file_handler)

# Also log to console in debug mode
if app.debug:
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter('%(levelname)s: %(message)s'))
    logger.addHandler(console_handler)

logger.info('Application startup')

# File access locks
events_file_lock = threading.Lock()
registrations_file_lock = threading.Lock()

# Input validation patterns
STUDENT_ID_PATTERN = re.compile(r'^\d{7}$')  # 7-digit student ID
EMAIL_PATTERN = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
NAME_PATTERN = re.compile(r'^[A-Za-z\s\'-]{2,50}$')

# Flask-Login initialization
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'admin.login'
login_manager.session_protection = 'strong'

# User model for Flask-Login
class User(UserMixin):
    def __init__(self, id, is_admin=False):
        self.id = id
        self.is_admin = is_admin

@login_manager.user_loader
def load_user(user_id):
    if user_id == ADMIN_USERNAME:
        return User(ADMIN_USERNAME, is_admin=True)
    return None

# Login attempt tracking with cleanup
class LoginAttemptTracker:
    def __init__(self, max_attempts, lockout_time):
        self.max_attempts = max_attempts
        self.lockout_time = lockout_time
        self.attempts = {}
        self._cleanup_interval = 3600  # Clean up every hour
        self._last_cleanup = time()

    def cleanup_old_attempts(self):
        current_time = time()
        if current_time - self._last_cleanup > self._cleanup_interval:
            self.attempts = {
                ip: data for ip, data in self.attempts.items()
                if current_time - data['last_attempt'] < self.lockout_time
            }
            self._last_cleanup = current_time

    def check_attempts(self, ip):
        self.cleanup_old_attempts()
        if ip in self.attempts:
            attempts = self.attempts[ip]
            if attempts['count'] >= self.max_attempts:
                if time() - attempts['last_attempt'] < self.lockout_time:
                    return False, int(self.lockout_time - (time() - attempts['last_attempt']))
                self.attempts.pop(ip)
        return True, 0

    def record_attempt(self, ip, success):
        self.cleanup_old_attempts()
        if success and ip in self.attempts:
            self.attempts.pop(ip)
        elif not success:
            if ip not in self.attempts:
                self.attempts[ip] = {'count': 1, 'last_attempt': time()}
            else:
                self.attempts[ip]['count'] += 1
                self.attempts[ip]['last_attempt'] = time()

# Initialize login attempt tracker
login_tracker = LoginAttemptTracker(MAX_ATTEMPTS, LOCKOUT_TIME)

# Create Excel files if they don't exist
def init_excel_files():
    # Create registrations file
    if not os.path.exists(REGISTRATIONS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Student ID', 'Full Name', 'Email', 'Event Name', 'Registration Date'])
        wb.save(REGISTRATIONS_FILE)
    
    # Create events file
    if not os.path.exists(EVENTS_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(['Event Name', 'Description', 'Date', 'Capacity', 'Status'])
        # Add some default events with future dates
        next_year = datetime.now().year + 1
        ws.append(['Orientation Day', 'Welcome event for new students', f'{next_year}-05-01', '100', 'Active'])
        ws.append(['Career Fair', 'Annual university career fair', f'{next_year}-06-15', '200', 'Active'])
        wb.save(EVENTS_FILE)
    else:
        # Update existing events to future dates if they're in the past
        try:
            wb = load_workbook(EVENTS_FILE)
            ws = wb.active
            next_year = datetime.now().year + 1
            current_date = datetime.now().date()
            file_updated = False

            for row in list(ws.iter_rows(min_row=2)):
                try:
                    event_date = datetime.strptime(row[2].value, '%Y-%m-%d').date()
                    if event_date < current_date:
                        # Update to next year's date
                        new_date = event_date.replace(year=next_year)
                        row[2].value = new_date.strftime('%Y-%m-%d')
                        file_updated = True
                except (ValueError, TypeError):
                    continue

            if file_updated:
                wb.save(EVENTS_FILE)
                logger.info('Events file updated with future dates')

        except Exception as e:
            logger.error(f"Error updating event dates: {e}")
            # If there's an error, create a new file
            wb = Workbook()
            ws = wb.active
            ws.append(['Event Name', 'Description', 'Date', 'Capacity', 'Status'])
            next_year = datetime.now().year + 1
            ws.append(['Orientation Day', 'Welcome event for new students', f'{next_year}-05-01', '100', 'Active'])
            ws.append(['Career Fair', 'Annual university career fair', f'{next_year}-06-15', '200', 'Active'])
            wb.save(EVENTS_FILE)

init_excel_files()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/events')
def view_events():
    wb = load_workbook(EVENTS_FILE)
    ws = wb.active
    events = []
    
    # Load registrations data for available slots
    try:
        wb_reg = load_workbook(REGISTRATIONS_FILE)
        ws_reg = wb_reg.active
        registration_counts = {}
        for row in ws_reg.iter_rows(min_row=2):
            event_name = row[3].value
            registration_counts[event_name] = registration_counts.get(event_name, 0) + 1
    except:
        registration_counts = {}
    
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        if row[4].value == 'Active':  # Only show active events
            event_name = row[0].value
            capacity = int(row[3].value)
            registered = registration_counts.get(event_name, 0)
            events.append({
                'id': idx,
                'name': event_name,
                'description': row[1].value,
                'date': row[2].value,
                'capacity': capacity,
                'available_slots': capacity - registered
            })
    return render_template('events.html', events=events)

@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("5 per minute")  # Rate limit registration attempts
def register():
    form = RegistrationForm()
    
    # Load available events for the form's select field
    try:
        with events_file_lock:
            wb = load_workbook(EVENTS_FILE)
            ws = wb.active
            events = []
            current_date = datetime.now().date()
            logger.info(f'Current date: {current_date}')
            
            # Populate events list for both form choices and display
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
                if row[4].value == 'Active':  # Only show active events
                    try:
                        event_date = datetime.strptime(row[2].value, '%Y-%m-%d').date()
                        logger.info(f'Processing event: {row[0].value}, Date: {event_date}, Status: {row[4].value}')
                        
                        if event_date >= current_date:  # Compare dates without time
                            current_registrations = 0
                            with registrations_file_lock:
                                wb_reg = load_workbook(REGISTRATIONS_FILE)
                                ws_reg = wb_reg.active
                                current_registrations = sum(1 for r in ws_reg.iter_rows(min_row=2) 
                                                         if r[3].value == row[0].value)
                            
                            logger.info(f'Adding event to list: {row[0].value}')
                            events.append({
                                'id': idx,
                                'name': row[0].value,
                                'description': row[1].value,
                                'date': event_date.strftime('%Y-%m-%d'),
                                'capacity': int(row[3].value),
                                'available_slots': int(row[3].value) - current_registrations,
                                'active': True
                            })
                    except (ValueError, TypeError) as e:
                        logger.error(f'Error processing event data: {str(e)}')
                        continue
            
            # Update form choices
            form.event.choices = [(e['id'], f"{e['name']} - {e['date']}") for e in events]
            
            if request.method == 'POST' and form.validate_on_submit():
                student_id = form.student_id.data
                name = form.name.data
                email = form.email.data
                event_id = form.event.data
                
                # Get event details and check capacity
                event_name = None
                event_capacity = None
                event_date = None
                
                for idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
                    if idx == event_id and row[4].value == 'Active':
                        event_name = row[0].value
                        event_capacity = int(row[3].value)
                        event_date = datetime.strptime(row[2].value, '%Y-%m-%d').date()
                        break
                
                if event_name is None:
                    flash('Selected event is not available', 'error')
                    return render_template('register.html', form=form, events=events)
                
                # Check if event date has passed
                if event_date < current_date:
                    flash('This event has already passed', 'error')
                    return render_template('register.html', form=form, events=events)
                
                # Check registration with file lock
                with registrations_file_lock:
                    wb_reg = load_workbook(REGISTRATIONS_FILE)
                    ws_reg = wb_reg.active
                    
                    # Check if the event is full
                    current_registrations = sum(1 for row in ws_reg.iter_rows(min_row=2) 
                                             if row[3].value == event_name)
                    
                    if current_registrations >= event_capacity:
                        flash('Sorry, this event is already full', 'error')
                        return render_template('register.html', form=form, events=events)
                    
                    # Check if student is already registered for this event
                    for row in ws_reg.iter_rows(min_row=2):
                        if row[0].value == student_id and row[3].value == event_name:
                            flash('You are already registered for this event', 'error')
                            return render_template('register.html', form=form, events=events)
                    
                    # Add registration
                    registration_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ws_reg.append([
                        student_id,
                        name,
                        email,
                        event_name,
                        registration_time
                    ])
                    wb_reg.save(REGISTRATIONS_FILE)
                    
                    # Log the successful registration
                    logger.info(f'New registration: Student {student_id} registered for {event_name}')
                    flash('Registration successful!', 'success')
                    return redirect(url_for('index'))
            
            logger.info(f'Total events found: {len(events)}')
            return render_template('register.html', form=form, events=events)
            
    except Exception as e:
        logger.error(f'Error loading registration page: {str(e)}')
        flash('Unable to load events. Please try again later.', 'error')
        return redirect(url_for('index'))

# Session timeout middleware
@app.before_request
def check_session_timeout():
    if current_user.is_authenticated:
        last_activity = session.get('last_activity', 0)
        if time() - last_activity > 1800:  # 30 minutes timeout
            session.clear()
            logout_user()
            flash('Session expired. Please log in again.', 'error')
            return redirect(url_for('admin.login'))
        session['last_activity'] = time()

# Security headers middleware
@app.after_request
def add_security_headers(response):
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' 'unsafe-eval'; style-src 'self' 'unsafe-inline';"
    return response

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('error.html', error='Page not found'), 404

@app.errorhandler(500)
def internal_error(error):
    logger.error(f'Server Error: {error}')
    return render_template('error.html', error='Internal server error'), 500

@app.errorhandler(429)
def too_many_requests(error):
    return render_template('error.html', error='Too many requests. Please try again later.'), 429

@app.route('/registrations')
@login_required  # Only logged-in admin can view registrations
def view_registrations():
    try:
        # Get filter parameter
        event_filter = request.args.get('event_filter', '')
        
        # Load all events for the filter dropdown
        with events_file_lock:
            wb_events = load_workbook(EVENTS_FILE)
            ws_events = wb_events.active
            events = [row[0].value for row in ws_events.iter_rows(min_row=2) if row[0].value]
        
        # Load registrations
        with registrations_file_lock:
            wb = load_workbook(REGISTRATIONS_FILE)
            ws = wb.active
            registrations = []
            
            # Skip header row and process registrations
            for row in ws.iter_rows(min_row=2):
                # Only include rows that match the filter, if one is applied
                if not event_filter or row[3].value == event_filter:
                    registrations.append({
                        'student_id': row[0].value,
                        'name': row[1].value,
                        'email': row[2].value,
                        'event': row[3].value,
                        'registration_date': row[4].value
                    })
            
            # Sort registrations by date (newest first)
            registrations.sort(key=lambda x: x['registration_date'], reverse=True)
        
        return render_template('view_registrations.html',
                             registrations=registrations,
                             events=sorted(events),
                             selected_event=event_filter)
    
    except Exception as e:
        logger.error(f'Error viewing registrations: {str(e)}')
        flash('Error loading registrations. Please try again later.', 'error')
        return redirect(url_for('admin.dashboard'))

if __name__ == '__main__':
    # Check if we're running in production mode
    is_production = os.environ.get('FLASK_ENV') == 'production'
    
    if is_production:
        # Production settings
        app.config['DEBUG'] = False
        app.config['TESTING'] = False
        app.config['PROPAGATE_EXCEPTIONS'] = True
        
        # Log to file only in production
        logger.info('Running in production mode')
        app.run(host='0.0.0.0', port=8080, ssl_context='adhoc')  # Enable HTTPS
    else:
        # Development settings
        app.config['DEBUG'] = True
        app.config['TESTING'] = True
        logger.info('Running in development mode')
        app.run(debug=True) 