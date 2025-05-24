from openpyxl import load_workbook

def check_events():
    wb = load_workbook('events.xlsx')
    ws = wb.active
    print('Events:')
    for row in ws.iter_rows(min_row=2):
        print(f'Name: {row[0].value}, Date: {row[2].value}, Status: {row[4].value}') 